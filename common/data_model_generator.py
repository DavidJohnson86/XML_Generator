"""Module responsibility to create JSON based datamodel which will be used
as an input format to create specific XML files."""

# pylint: disable=F0401
# disable unable to import error

import os
import json
from enum import Enum
from common import config
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet

TEST_DATA_PATH = r'\projects\AMVH422ev.xlsx'


class Mandatory(Enum):
    """Contains XML tags what is mandatory for all test cases"""
    REPORT_NAME = {config.REPORT_NAME_TAG: config.REPORT_NAME}
    CAN_SIG = {config.SIGNAL_MONITOR_TAG: config.SIGNAL_MONITOR}
    CRASH_LIST = {config.CRASH_LIST_TAG: config.CRASH_LIST}
    ESP_V_REF_VAL = {config.ESP_V_REF_VAL_TAG : config.ESP_V_REF_VAL}


class SquibFaultHandling(Enum):
    """Contains fault types all test case types"""
    OPEN = "OPEN"
    SHORT = "SHORT"
    LEAK_BAT = "LEAK_BAT"
    LEAK_GND = "LEAK_GND"


class TestGenerator:
    """This Class is responsible for create JSON format from test datas
    It creates testcases in JSON format that data could be passed to the
    variation generator what creates XML files for test input data"""

    # pylint: disable=R0903
    # disable too few public methods

    def __init__(self, component_data, error_set):
        """
        Initialize variables

        Args:
            component_data (list of dicts): Component information
            error_set(enum): Enum class variables of fault types
        """
        self.component_data = component_data
        self.error_set = error_set
        self.data = {}

    def add_tags(self):
        """
        Creates test cases for each error_set for each component defined in component_data
        and also adds mandatory elements.

        """
        for component in self.component_data:
            for test in self.error_set:
                test_case_name = "{0}_{1}".format(component["Name"], test.value)
                self.data[test_case_name] = []

                # Add Mandatory Tags
                for enum in Mandatory:
                    self.data[test_case_name].append(enum.value)

                # Add Switch States
                self.data[test_case_name].append(
                    {config.SWITCHING: "{0}:{1}".format(component["Switch"], test.value)})

                # Add DID Tags
                self.data[test_case_name].append({config.READ_DID_TAG: component["DID"]})
        return self.data


def parse_testcases(path, sheet_name):
    """
    Parse test case information from an input file.

    Args:
        sheet_name (str): sheet name

    Examples:
        # >>> parse_testcases("sheet")
    """

    # get absolute path of the the file
    working_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    path = working_dir + path

    # open excel sheet
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    assert isinstance(sheet, Worksheet)

    # read content
    testcases = []
    row_idx = 1
    keys = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

        # collect keys from the header
        if row_idx == 1:
            keys = [str(cell.value) for cell in row]

        else:
            values = [str(cell.value) for cell in row]
            test_case = dict(zip(keys, values))
            testcases.append(test_case)
        row_idx += 1
    return testcases


if __name__ == "__main__":
    COMPONENT_DATA = parse_testcases(TEST_DATA_PATH, "squibs")
    DATA = TestGenerator(COMPONENT_DATA, SquibFaultHandling).add_tags()
    print(json.dumps(DATA, indent=4))
