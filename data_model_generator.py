import config
from enum import Enum
import json
import os
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet

TEST_DATA_PATH = r'\test_data.xlsx'



class Mandatory(Enum):
    """Contains XML tags what is mandatory for all test cases"""
    REPORT_NAME = {config.REPORT_NAME_TAG: config.REPORT_NAME}
    CAN_SIG = {config.SIGNAL_MONITOR_TAG: config.SIGNAL_MONITOR}
    CRASH_LIST = {config.CRASH_LIST_TAG: config.CRASH_LIST}
    ESP_V_REF_VAL = {config.ESP_V_REF_VAL_TAG : config.ESP_V_REF_VAL}


class SquibFaultHandling(Enum):
    """Contains XML tags what is mandatory for all test cases"""
    OPEN = "OPEN"
    SHORT = "SHORT"
    LEAK_BAT = "LEAK_BAT"
    LEAK_GND = "LEAK_GND"


class TestGenerator:
    """This Class is responsible for create JSON format from test datas
    It creates testcases in JSON format that data could be passed to the
    variation generator what creates XML files for test input data"""

    def __init__(self, test_cases, component):
        """
        Initialize variables

        Args:
            test_cases (enum object): Enum Class for test case definition
        """
        self.test_cases = test_cases
        self.component = component
        self.scenario = None
        self.data = {}

    def add_tags(self):
        """
        Creates test cases for each component for each scenario defined in test_cases
        and also adds mandatory elements.

        """
        for self.scenario in self.test_cases:
            for test in self.component:
                test_case_name = "{0}_{1}".format(self.scenario["Name"], test.value)
                self.data[test_case_name] = []

                # Add Mandatory Tags
                for enum in Mandatory:
                    self.data[test_case_name].append(enum.value)

                # Add Switch States
                self.data[test_case_name].append(
                    {config.SWITCHING: "{0}:{1}".format(self.scenario["Switch"], test.value)})

                # Add DID Tags
                self.data[test_case_name].append({config.READ_DID_TAG: self.scenario["DID"]})
        return self.data


def parse_testcases(path, sheet_name):
    """
    Parse test case information from an input file.

    Args:
        sheet_name (str): sheet name

    Examples:
        # >>> parse_testcases("sat1")
    """

    # get absolute path of the real-time sequence
    working_dir = os.path.dirname(os.path.realpath(__file__))
    path = working_dir + path

    # open excel sheet
    workbook = load_workbook(path)
    sheet = workbook[sheet_name]
    assert isinstance(sheet, Worksheet)

    # read content
    testcases = []
    row_idx = 1
    keys = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

        # collect keys from the header
        if row_idx == 1:
            keys = [str(cell.value) for cell in row]

        else:
            values = [str(cell.value) for cell in row]
            test_case = dict(zip(keys, values))
            testcases.append(test_case)
        row_idx += 1
    return testcases


if __name__ == "__main__":
    test_inputs = parse_testcases(TEST_DATA_PATH, "squibs")
    data = TestGenerator(test_inputs, SquibFaultHandling).add_tags()
    print(json.dumps(data, indent=4))


